import pandas as pd 
import numpy as np 
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import streamlit as st 


class StylingReports:
    def _is_blank(self, df: pd.DataFrame):
        """ Standarize Missing Value, Convert Missing Values to Format Excel"""
        try : 
            if df is None:
                return True
            
            if isinstance(df, float) and math.isnan(df):
                return True
            
            if isinstance(df, str) and df.strip() == "":
                return True
            return False
        
        except Exception as e :
            st.error(f"Error in class StylingReports (_is_blank) : {e}")
    
    def _write_df_to_worksheet(self, df: pd.DataFrame):
        """ Import Data to Library Openpyxl """
        try : 
            df = df.where(pd.notna(df), None)
            
            # Optional: bersihkan nilai aneh
            df = df.applymap(lambda x: str(x) if isinstance(x, (dict, list, tuple)) else x)

            wb = Workbook()
            ws = wb.active
      
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            print(f"Done Import")
                
            return wb, ws
        
        except Exception as e : 
            st.error(f"Error in class StylingReports (_write_df_to_worksheet) : {e}")
            return None, None
    
    def _merge_cells(self, ws, start_col, end_col, row_start=1, max_row=None):
        try : 
            align = Alignment(horizontal="center", vertical="center")
            max_row = max_row

            for col in range(start_col, end_col):
                row = row_start
                while row <= max_row:
                    top_val = ws.cell(row, col).value
                    if self._is_blank(top_val):
                        row += 1
                        continue

                    end = row
                    while end + 1 <= max_row and self._is_blank(ws.cell(end + 1, col).value):
                        ws.cell(end + 1, col).value = top_val
                        end += 1
                    
                    if end > row : 
                        ws.merge_cells(start_row=row, start_column=col, end_row=end, end_column=col)
                        ws.cell(row, col).alignment = align
                    row = end + 1
            
            #self._apply_center_alignment(ws) 
        
        except Exception as e : 
            st.error(f"Error in class StylingReports (_merge_cells) : {e}")
    
    def run(self, df, start_col1, start_col2, end_col1, end_col2, option_val): 
        """ 
            Generate Styling Reports with openpyxl.

            parameters : 
                - df : DataFrame will be add styling.
                - start_col1 : The column to be styling (column start for header)
                - start_col2 : The column to be styling (column start for values)
                - end_col1 : The column to be styling (column end for header)
                - end_col2 : The column to be styling (column end for values)
                - row_start : Styling started with row number (if 1, started styling in row number 1)
                - option : Format to Generate Reports 
        """
        try : 
            wb, ws = self._write_df_to_worksheet(df)
            
            if option_val == "Format 1" : 
                self._merge_cells(ws, start_col1, end_col1, row_start=2, max_row=ws.max_row)
                self._merge_cells(ws, start_col2, end_col2, row_start=2, max_row=ws.max_row)  
                
                return wb  
            
            elif option_val == "Format 2" :  
                self._merge_cells(ws, start_col1, end_col1, row_start=1, max_row=4)
                self._merge_cells(ws, start_col2, end_col2, row_start=1, max_row=4)

                self._merge_cells(ws, start_col1, end_col1, row_start=2, max_row=ws.max_row)
                self._merge_cells(ws, start_col2, end_col2, row_start=2, max_row=ws.max_row)

                return wb  
        
        except Exception as e : 
            st.error(f"Error in class StylingReports (run) : {e}")